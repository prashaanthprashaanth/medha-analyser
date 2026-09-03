#!/usr/bin/env python3
"""Local HTTP bridge used by the Electron Medha desktop application."""

from __future__ import annotations

import argparse
import base64
import ctypes
import hmac
import json
import mimetypes
import os
import secrets
import shutil
import sys
import tempfile
import threading
import time
import webbrowser
from io import BytesIO
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urlsplit

from medha_analyser import FormatError
from medha_service import AnalysisService


def parent_is_running(process_id: int) -> bool:
    """Check the Electron parent without requiring third-party packages."""

    if process_id <= 0:
        return True
    if os.name == "nt":
        process_query_limited_information = 0x1000
        still_active = 259
        handle = ctypes.windll.kernel32.OpenProcess(
            process_query_limited_information, False, process_id
        )
        if not handle:
            return False
        try:
            exit_code = ctypes.c_ulong()
            return bool(
                ctypes.windll.kernel32.GetExitCodeProcess(handle, ctypes.byref(exit_code))
                and exit_code.value == still_active
            )
        finally:
            ctypes.windll.kernel32.CloseHandle(handle)
    try:
        os.kill(process_id, 0)
    except OSError:
        return False
    return True


def install_lifetime_watchers(
    server: ThreadingHTTPServer, parent_pid: int
) -> threading.Event:
    """Stop the local API when Electron closes or its stdin pipe disappears."""

    stopping = threading.Event()

    def request_shutdown() -> None:
        if stopping.is_set():
            return
        stopping.set()
        server.shutdown()

    def watch_parent() -> None:
        while not stopping.wait(1.0):
            if parent_pid and not parent_is_running(parent_pid):
                request_shutdown()
                return

    def watch_pipe() -> None:
        try:
            stream = getattr(sys.stdin, "buffer", sys.stdin)
            if stream is not None:
                while stream.read(1):
                    pass
                request_shutdown()
        except (OSError, ValueError):
            request_shutdown()

    if parent_pid:
        threading.Thread(target=watch_parent, name="electron-watch", daemon=True).start()
        threading.Thread(target=watch_pipe, name="electron-pipe", daemon=True).start()
    return stopping


def make_excel(payload: dict[str, Any]) -> dict[str, str]:
    """Create a styled XLSX workbook for desktop renderer exports."""

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    headers = list(payload.get("headers") or [])
    rows = list(payload.get("rows") or [])
    if not headers:
        raise ValueError("Excel export requires column headings")
    if len(headers) * (len(rows) + 1) > 600_000:
        raise ValueError("Excel export is too large; reduce the selected time range")
    workbook = Workbook()
    sheet = workbook.active
    title = str(payload.get("sheet") or "Medha Data")
    for character in "[]:*?/\\":
        title = title.replace(character, "_")
    sheet.title = title[:31] or "Medha Data"
    sheet.append(headers)
    for row in rows:
        sheet.append(list(row))
    fill = PatternFill("solid", fgColor="0C4F78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, header in enumerate(headers, start=1):
        observed = [len(str(header))]
        for row in rows[:250]:
            if index - 1 < len(row):
                observed.append(len(str(row[index - 1] if row[index - 1] is not None else "")))
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = min(55, max(10, max(observed) + 2))
    output = BytesIO()
    workbook.save(output)
    return {"base64": base64.b64encode(output.getvalue()).decode("ascii")}


class ApiHandler(BaseHTTPRequestHandler):
    service: AnalysisService
    quiet = False
    browser_token = ""
    browser_upload_dir: Path | None = None
    browser_last_heartbeat = 0.0
    browser_heartbeat_started = False
    browser_busy_until = 0.0
    browser_lock = threading.Lock()

    @classmethod
    def note_browser_heartbeat(cls) -> None:
        with cls.browser_lock:
            cls.browser_last_heartbeat = time.monotonic()
            cls.browser_heartbeat_started = True
            cls.browser_busy_until = 0.0

    @classmethod
    def note_browser_busy(cls) -> None:
        with cls.browser_lock:
            cls.browser_last_heartbeat = time.monotonic()
            cls.browser_heartbeat_started = True
            cls.browser_busy_until = time.monotonic() + 3600.0

    @classmethod
    def note_browser_closing(cls) -> None:
        with cls.browser_lock:
            cls.browser_heartbeat_started = True
            cls.browser_busy_until = 0.0
            cls.browser_last_heartbeat = time.monotonic() - 9.0

    def _authorized(self) -> bool:
        if not self.browser_token:
            return True
        supplied = self.headers.get("X-Medha-Token", "")
        return hmac.compare_digest(supplied, self.browser_token)

    def _static_root(self) -> Path:
        bundle_root = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
        packaged = bundle_root / "viewer"
        return packaged if packaged.is_dir() else Path(__file__).resolve().parent / "medha_desktop" / "viewer"

    def _serve_static(self, request_path: str) -> bool:
        routes = {
            "/": "index.html",
            "/index.html": "index.html",
            "/detail.html": "detail.html",
            "/styles.css": "styles.css",
            "/history_chart.js": "history_chart.js",
            "/browser_bridge.js": "browser_bridge.js",
            "/app.js": "app.js",
            "/detail.js": "detail.js",
        }
        filename = routes.get(request_path)
        if filename is None:
            return False
        source = self._static_root() / filename
        if not source.is_file():
            self._send(404, {"error": f"Application resource is missing: {filename}"})
            return True
        data = source.read_bytes()
        content_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
        self.send_response(200)
        self.send_header("Content-Type", content_type + ("; charset=utf-8" if content_type.startswith("text/") or content_type.endswith("javascript") else ""))
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.end_headers()
        self.wfile.write(data)
        return True

    def _receive_archive(self) -> dict[str, str]:
        if self.browser_upload_dir is None:
            raise ValueError("Browser uploads are not enabled")
        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0:
            raise ValueError("Select a locomotive ALL-data ZIP")
        if length > 8 * 1024 * 1024 * 1024:
            raise ValueError("ZIP is larger than the supported 8 GB limit")
        original = Path(unquote(self.headers.get("X-Medha-Filename", "locomotive-all-data.zip"))).name
        if not original.lower().endswith(".zip"):
            raise ValueError("Select the original locomotive ALL-data ZIP")
        destination = self.browser_upload_dir / f"{secrets.token_hex(6)}-{original}"
        remaining = length
        with destination.open("wb") as output:
            while remaining:
                chunk = self.rfile.read(min(1024 * 1024, remaining))
                if not chunk:
                    raise ValueError("The ZIP upload ended before it was complete")
                output.write(chunk)
                remaining -= len(chunk)
        return {"path": str(destination), "name": original}

    def log_message(self, message: str, *args: object) -> None:
        if not self.quiet and sys.stderr is not None:
            print(message % args, file=sys.stderr)

    def _json_body(self) -> dict[str, Any]:
        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0:
            return {}
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def _send(self, status: int, value: object) -> None:
        data = json.dumps(value, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self) -> None:  # noqa: N802 - BaseHTTPRequestHandler API
        path = urlsplit(self.path).path
        try:
            if self._serve_static(path):
                return
            if not self._authorized():
                self._send(403, {"error": "This analyser session is no longer active"})
            elif path == "/health":
                self._send(200, {"ok": True})
            elif path == "/status":
                self._send(200, self.service.status())
            elif path == "/faults":
                self._send(200, {"faults": self.service.get_faults(), "status": self.service.status()})
            elif path == "/fault-parameters":
                self._send(200, self.service.fault_parameters())
            else:
                self._send(404, {"error": "Unknown endpoint"})
        except Exception as exc:
            self._send(500, {"error": str(exc)})

    def do_POST(self) -> None:  # noqa: N802 - BaseHTTPRequestHandler API
        path = urlsplit(self.path).path
        try:
            if not self._authorized():
                self._send(403, {"error": "This analyser session is no longer active"})
                return
            if path == "/heartbeat":
                self.note_browser_heartbeat()
                self._send(200, {"ok": True})
                return
            if path == "/browser-busy":
                self.note_browser_busy()
                self._send(200, {"ok": True})
                return
            if path == "/browser-closing":
                self.note_browser_closing()
                self._send(200, {"ok": True})
                return
            if path == "/upload":
                self._send(200, self._receive_archive())
                return
            body = self._json_body()
            if path == "/load":
                result = self.service.load_archive(body.get("path", ""))
            elif path == "/fault-detail":
                result = self.service.fault_detail(
                    int(body["row_index"]), int(body.get("snapshot_choice", 0))
                )
            elif path == "/fault-comparison":
                result = self.service.fault_comparison(
                    body.get("row_indices", []), body.get("parameters", [])
                )
            elif path == "/history":
                result = self.service.history_page(
                    body["memory"],
                    body.get("parameters"),
                    body.get("start"),
                    body.get("end"),
                    body.get("offset", 0),
                    body.get("limit", 1000),
                    body.get("newest_first", True),
                )
            elif path == "/history-chart":
                result = self.service.history_chart(
                    body["memory"],
                    body.get("parameters", []),
                    body.get("start"),
                    body.get("end"),
                    body.get("max_points", 12000),
                )
            elif path == "/make-excel":
                result = make_excel(body)
            else:
                self._send(404, {"error": "Unknown endpoint"})
                return
            self._send(200, result)
        except (FormatError, KeyError, TypeError, ValueError, OSError) as exc:
            self._send(400, {"error": str(exc)})
        except Exception as exc:
            self._send(500, {"error": str(exc)})


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=0)
    parser.add_argument("--parent-pid", type=int, default=0)
    parser.add_argument("--browser", action="store_true")
    parser.add_argument("--no-open-browser", action="store_true", help=argparse.SUPPRESS)
    parser.add_argument("--session-token", default="", help=argparse.SUPPRESS)
    parser.add_argument("--quiet", action="store_true", help=argparse.SUPPRESS)
    parser.add_argument("--ready-file", default="", help=argparse.SUPPRESS)
    parser.add_argument("--version", action="version", version="Medha Data Analyser 1.0")
    args = parser.parse_args()
    browser_mode = bool(args.browser or (getattr(sys, "frozen", False) and not args.parent_pid))
    upload_directory: tempfile.TemporaryDirectory[str] | None = None
    service = AnalysisService()
    ApiHandler.service = service
    ApiHandler.quiet = bool(args.quiet or (browser_mode and getattr(sys, "frozen", False)))
    if browser_mode:
        upload_root = Path(tempfile.gettempdir())
        cutoff = time.time() - 6 * 60 * 60
        for old_upload in upload_root.glob("medha-analyser-*"):
            try:
                if old_upload.is_dir() and old_upload.stat().st_mtime < cutoff:
                    shutil.rmtree(old_upload, ignore_errors=True)
            except OSError:
                pass
        upload_directory = tempfile.TemporaryDirectory(prefix="medha-analyser-")
        ApiHandler.browser_token = args.session_token or secrets.token_urlsafe(24)
        ApiHandler.browser_upload_dir = Path(upload_directory.name)
        ApiHandler.browser_last_heartbeat = time.monotonic()
        ApiHandler.browser_heartbeat_started = False
        ApiHandler.browser_busy_until = 0.0
    server = ThreadingHTTPServer((args.host, args.port), ApiHandler)
    stopping = install_lifetime_watchers(server, args.parent_pid)
    host, port = server.server_address
    application_url = (
        f"http://{host}:{port}/?session={ApiHandler.browser_token}" if browser_mode else None
    )
    ready_message = {"ready": True, "host": host, "port": port, "url": application_url}
    if args.ready_file:
        Path(args.ready_file).write_text(json.dumps(ready_message), encoding="utf-8")
    if not args.quiet and sys.stdout is not None and (
        not getattr(sys, "frozen", False) or args.parent_pid
    ):
        print(json.dumps(ready_message), flush=True)

    if browser_mode:
        def open_application() -> None:
            time.sleep(0.45)
            webbrowser.open(application_url, new=2)

        def watch_browser() -> None:
            started = time.monotonic()
            while not stopping.wait(1.0):
                with ApiHandler.browser_lock:
                    heartbeat_started = ApiHandler.browser_heartbeat_started
                    heartbeat_age = time.monotonic() - ApiHandler.browser_last_heartbeat
                    busy = time.monotonic() < ApiHandler.browser_busy_until
                if busy:
                    continue
                if (heartbeat_started and heartbeat_age > 8.0) or (
                    not heartbeat_started and time.monotonic() - started > 60.0
                ):
                    stopping.set()
                    server.shutdown()
                    return

        if not args.no_open_browser:
            threading.Thread(target=open_application, name="browser-open", daemon=True).start()
        threading.Thread(target=watch_browser, name="browser-watch", daemon=True).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        if browser_mode:
            # Decoder workers are read-only but may still be indexing a large ZIP.
            # Guarantee prompt process exit even if Windows still holds an upload handle.
            force_exit = threading.Timer(1.0, lambda: os._exit(0))
            force_exit.daemon = True
            force_exit.start()
        stopping.set()
        server.server_close()
        service.executor.shutdown(wait=False, cancel_futures=True)
        if service.loco is not None:
            service.loco.close()
        if upload_directory is not None:
            try:
                upload_directory.cleanup()
            except OSError:
                pass
        if browser_mode:
            # Bypass Python's executor atexit join after the read-only workers
            # have been cancelled; the next launch always starts a fresh session.
            os._exit(0)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        if getattr(sys, "frozen", False) and os.name == "nt":
            ctypes.windll.user32.MessageBoxW(
                0,
                f"Medha Data Analyser could not start.\n\n{error}",
                "Medha Data Analyser 1.0",
                0x10,
            )
        raise
